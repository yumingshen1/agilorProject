# -*- coding: utf-8 -*-
# @Time: 2022/2/23 11:32
# @Author: shenyuming

'''
    教程
'''
from openpyxl import Workbook
from openpyxl import load_workbook
#创建工作簿
wb = Workbook()
# 获取工作表
ws = wb.active
# 创建新的工作表
ws1 = wb.create_sheet('Mysheet')
ws2 = wb.create_sheet('Mysheet2',0)
ws3 = wb.create_sheet('Mysheet3',-1)

# 修改工作表名字
ws.title = 'New title'
# 更改工作表颜色
ws.sheet_properties.tabColor = '1072BA'
# 将修改的工作表做为工作簿的键
ws = wb['New title']
## 查看只能所有的工作表名称
print('工作表',wb.sheetnames)
## 循环打印工作表
for i in wb:
    print('循环表名:',i.title)

## 创建ws的工作表副本
ws4 = wb.copy_worksheet(ws)
ws4.title = '复制的New title'
print('工作表',wb.sheetnames)

##  数据，访问一个单元格
C = ws['A4']
ws['A4'] = 4
#指定单元格写入
d = ws.cell(row=4,column=2,value=10)
print('指定单元格d的数据：',d)
## 创建100x100单元格
for x in range(1,101):
    for y in range(1,101):
        ws.cell(row=x,column=y)

wb.save('../excelData/blance.xlsx')

## 访问多个单元格
cell_range = ws['A1':'C1']
print('访问的多个单元格：',cell_range)
    ## min_row---从第几行开始，   max_col---最大的列数，  max_row---最大行数
for row in ws.iter_rows(min_row=1,max_col=3,max_row=2):
    for cell in row:
        print('cell行的值：',cell)
for col in ws.iter_cols(min_row=1,max_col=3,max_row=2):
    for cell in col:
        print('cell列的值',cell)

# print('全部行：',tuple(ws.rows))
# print('全部列：',tuple(ws.cols))

## 遍历工作表每一行单元格的值
for row in ws.values:
    for value in row:
        pass
    # print('单元格的值：',value)
for i in ws.iter_rows(min_row=1,max_col=3,max_row=4,values_only=True):
    print('遍历单元格的数据：',i)


## 打开现有excel
wb2 = load_workbook('../excelData/blance.xlsx')
print('查看有几个sheet:',wb2.sheetnames)