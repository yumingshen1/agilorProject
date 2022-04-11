# -*- coding:utf-8 -*-
# @Time : 2022/2/27 15:42
# Auther : shenyuming
# @File : openpy_01xl.py
# @Software : PyCharm

import openpyxl
import datetime
## 创建一个excel保存
path = f'/Users/shenyuming/Downloads/sym/openpyxl_excel/excel01.xlsx'
wb = openpyxl.Workbook()
ws = wb.active
ws['A1'] = 42
ws.append([1, 2, 3])
ws['A3'] = datetime.datetime.now()
wb.save(path)