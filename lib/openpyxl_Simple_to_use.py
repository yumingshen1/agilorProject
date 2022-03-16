# -*- coding: utf-8 -*-
# @Time: 2022/2/24 11:52
# @Author: shenyuming

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

'''
    写入新的工作簿
'''
wb = Workbook()

dest_filename = '../excelData/empty_book.xlsx'
ws1 = wb.active
ws1.title = 'range names'
for row in range(1,40):     # 行，30行
    ws1.append(range(600))  # 列，599列， 每列写入对应列的数字

ws2 = wb.create_sheet(title='PI')   # 创建新的工作表，
ws2['F5'] = 3.14                    # 单元格赋值

ws3 = wb.create_sheet(title='Data')     #创建第三个工作表
for row in range(10,20):                #行，第10行到第19行
    for col in range(27,54):            #列，第27列到第53列
        _ = ws3.cell(column=col,row=row,value="{0}".format(get_column_letter(col)))   # 写入对应列的字母
print('ws3的数据：',ws3['AA10'].value)   # 打印单元格值
for row in ws3.values:          # 循环打印excel的所有单元格值
    for col in row:
        # print('循环打印excel的所有值',col)
        pass
wb.save(filename=dest_filename)

'''
    打开现有工作簿
'''
from openpyxl import load_workbook
wb = load_workbook(filename='../excelData/empty_book.xlsx')
sheet_ranges = wb['range names']
print(sheet_ranges['D36'].value)


'''
    新工作簿写入
    使用数字格式
'''
import datetime
wb = Workbook()
ws = wb.active
ws['A1'] = datetime.datetime(2020,7,20)
print(ws['A1'].number_format)

'''
    使用公式
'''
ws['B1'] = "=SUM(1,1)"
# wb.save('../excelData/formula.xlsx')

'''
    插入图片
'''
from openpyxl.drawing.image import Image
image = Image('../excelData/代码01.png')
ws.add_image(image,'C1')
# wb.save('../excelData/formula.xlsx')


'''
    二维面积图
'''
from openpyxl.chart import (
    AreaChart,
    AreaChart3D,
    Reference,
    Series,
)
#创建新的sheet用于图标
ws2 = wb.create_sheet(title='tubiao')

rows = [
    ['Number', 'Batch 1', 'Batch 2'],
    [2, 40, 30],
    [3, 40, 25],
    [4, 50, 30],
    [5, 30, 10],
    [6, 25, 5],
    [7, 50, 10],
]

for row in rows:
    ws2.append(row)

chart = AreaChart()
chart.title = "Area Chart"
chart.style = 13
chart.x_axis.title = 'Test'
chart.y_axis.title = 'Percentage'

cats = Reference(ws2, min_col=1, min_row=1, max_row=7)
data = Reference(ws2, min_col=2, min_row=1, max_col=3, max_row=7)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
ws2.add_chart(chart, "A10")

wb.save('../excelData/formula.xlsx')