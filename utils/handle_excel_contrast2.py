# -*- coding: utf-8 -*-
# @Time: 2022/1/26 11:17
# @Author: shenyuming
#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
    excel大量数据对比, 对比多列数据
"""

#导入模块 openpyxl
# !/usr/bin/env python
# -*- coding: utf-8 -*-
# 导入模块 openpyxl
import operator
import openpyxl,xlsxwriter
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
import copy

class conTrastExcel:
    def __init__(self):
        # 读取excel文件
        # 括号中的字符串为你要比较的两个excel的路径，注意用“/”
        wb_a = openpyxl.load_workbook(f'E:/sym/PI/pi_424_data/pi_CDM158.xlsx')  #pi
        wb_b = openpyxl.load_workbook(f'E:/sym/PI/rtdb_424_421/rtdb_CDM158.xlsx')  #4.2
        self.wa = wb_a.worksheets[0]
        self.wb = wb_b.worksheets[0]

    print('-------第一张表---------')
    def excel1(self):
        lista = []
        for i in self.wa.iter_rows():
            list = []
            for cell in i:
                a = cell.value
                list.append(a)
            lista.append(list)
        return lista

    print('-------第二张表---------')
    def excel2(self):
        listb = []
        for j in self.wb.iter_rows():
            list = []
            for cell in j:
                b = cell.value
                list.append(b)
            listb.append(list)
        return listb
    # 对比
    def con(self,a,b,file_path):
        if len(a) > len(b):     # 第一个excel数据是否大于第二个execl数据，取list的长度
            print("---len(a)值：{}".format(len(a)))
            c = len(a)
            d = len(b)
        else:                   # 如果第一个excel长度小于或等于第二个excel长度，取list长度
            print("---len(b)值：{}".format(len(b)))
            c = len(b)             # 获取list长度，只为循环准备
            d = len(a)

        datas = []
        # c 为最大值；d 为最小值
        for i in range(c):          # 循环list长度
            if i + 1 <= d:  ## 判断 i+1 是否小于等于最短list的长度，
                tempList = []
                for j in range(len(a[i])):  ## 循环 a下标 长度，列数
                    if a[i][j] != b[i][j]:      ## 如果两值不等，
                        # wb_s.font = Font(color=colors.BLACK, italic=True, bold=True)
                        # wb_s.fill = PatternFill("solid", fgColor="DDDDDD")
                        tempList.append("{} 和 {} 不一致".format(a[i][j], b[i][j]))
                    else:
                        tempList.append(a[i][j])    # 两值相等直接加入list
                datas.append(tempList)      # 将每一组的数据，放入datas
            else:       # 如果i+1>d 最短list的长度， 并且第一个excel数据大于第二个excel数据，直接写入第一个excel的数据
                if len(a) > len(b):
                    datas.append(a[i])
                else:
                    datas.append(b[i])      # 如果i+1>d 最短list的长度， 并且第二个excel数据大于等于第一个excel数据 ，直接写入第二个excel数据
        print("---datas值：{}".format(datas))

        ## 创建工作簿，写入对比结果
        wb_s = Workbook()
        ws = wb_s.active
        ws.title = '对比结果'

        for index, item in enumerate(datas):
            print("---item： {}".format(item))
            ws.cell(index+1, 1, '{}'.format(item[0]))
            ws.cell(index+1, 2, '{}'.format(item[1]))
            ws.cell(index+1, 3, '{}'.format(item[2]))
            ws.cell(index+1, 4, '{}'.format(item[3]))
            ws.cell(index+1, 5, '{}'.format(item[4]))

        wb_s.save(file_path)


a = conTrastExcel().excel1()
b = conTrastExcel().excel2()
file_path = f'E:/sym/PI/对比结果pi_2022_425/CDM158_421.xlsx'
conTrastExcel().con(a,b,file_path)
