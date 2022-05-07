import requests
import pandas as pd
import json
import time
import openpyxl
from itertools import groupby

s_value_list = []  # 所有用户点的属性值列表，每个子列表是每个用户点的属性值[[val1,val2···], [val3,val4···], [val5,val6···]···]
column_val = []    # 属性name列表
name_val = []      # 所有用户点的Name列表

def request_data(url):
    r = requests.request("get", url)
    #time.sleep(900)
    data = json.loads(r.text)

    Name_list = []
    WebId_list = []
    for i in range(len(data["Items"])):
        WebId_list.append(data["Items"][i]['WebId'])
        if "PI Batch Database generated, do not delete or edit." not in data["Items"][i]['Descriptor']:  # 判断是否是系统点
            Name_list.append(data["Items"][i]['Name'].replace(" ", ""))
    #print("PointName列表：", Name_list)
    #print(len(Name_list))
    for item in Name_list:
        name_val.append(item)

    b_list = []  # 所有点的属性name,value
    for j in range(len(WebId_list)):
        api = 'http://192.168.30.72:8080/piwebapi/points/'
        url1 = api + WebId_list[j] + '/attributes'
        a = requests.request("get", url1)
        data1 = json.loads(a.text)
        for n in range(len(data1["Items"])):
            b_list.append(data1["Items"][n]['Name'])
            b_list.append(data1["Items"][n]['Value'])
    #print("总属性值列表：", b_list)
    b_list1 = []   # 转成列表中每个元素是字符串
    for i in b_list:
        b_list1.append(str(i))

    # 把所有点的属性name,value切割，每个子列表存储单个点的name,value
    i = (list(g) for _, g in groupby(b_list1, key='archiving'.__ne__))
    b_list2 = [a + b for a, b in zip(i, i)]
    #print("切割后属性值列表：", b_list2)

    # 排除系统点
    b_list_new = []    # 只有用户点name,value的大列表
    for item in b_list2:
        if len(item) == 116:
            b_list_new.append(item)
    #print("只有用户点的列表：", b_list_new)
    #print(len(b_list_new))

    for each_item in b_list_new:
        if isinstance(each_item, list):
            #s_name = each_item[::2]    # 每个子列表切片得到属性name
            s_value = each_item[1::2]   # 每个子列表切片得到属性value
            s_value_list.append(s_value)
    #print("哈哈哈哈：", s_value_list)

    #print("第一个点的属性：", b_list_new[0])
    s1_name = b_list_new[0][::2]  # 列表切片得到属性name（每个点的属性name都相同，所以这里取第一个点）
    for item in s1_name:
        column_val.append(item)
    #print("第一个点的属性name:", s1_name)
    #s1_value = b_list_new[0][1::2]
    #print("第一个点的属性value:", s1_value)

#excel写入
# def excel_write():
#     i = 0
#     for item in s_value_list:
#         item.insert(0, name_val[i])
#         i = i+1
#     column_val.insert(0, 'Name')
#     df = pd.DataFrame(s_value_list, columns=column_val)
#     df.to_excel('./db4_pi_418.xlsx', index=False)


def writeToExcel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(1, 1).value = 'PointName'
    for r in range(len(s_value_list)):
        for c in range(len(s_value_list[0])):
            ws.cell(r + 2, c + 2).value = s_value_list[r][c]
            # excel中的行和列是从1开始计数的，所以需要+1
    #属性名设置成表头
    for i in range(len(column_val)):
        ws.cell(1, i + 2).value = column_val[i]
    #用户点名设置成第一列
    for j in range(len(name_val)):
        ws.cell(j + 2, 1).value = name_val[j]
    wb.save(f'E:/sym/PI/pi_505_data/pi_attribute_43374.xlsx')
    return 1

#ss = [['a', 'b', 'c'], ['aa', 'bb', 'cc'], ['aaa', 'bbb', 'ccc']]
# def writeTotxt():
#     with open("./data.txt", "w") as f:
#         for i in s_value_list:
#             i = str(i).strip('[').strip(']').replace(', ', '@').replace('\'', '') + '\n'
#             f.write(i)

#列表拆分
def list_of_groups(init_list, children_list_len):
    list_of_groups = zip(*(iter(init_list),) *children_list_len)
    end_list = [list(i) for i in list_of_groups]
    count = len(init_list) % children_list_len
    end_list.append(init_list[-count:]) if count !=0 else end_list
    return end_list

if __name__ == '__main__':
    url = 'http://192.168.30.72:8080/piwebapi/dataservers/F1DS3uSn5IfY2kGMucN6_OSrNAV0lOLVEzNzRQUEdBSDZD/points?maxCount=43376'
    request_data(url)
    writeToExcel()
