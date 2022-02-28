# -*- coding: utf-8 -*-
# @Time: 2022/1/21 14:22
# @Author: shenyuming

import requests,re,openpyxl
import json,time,datetime
from openpyxl.styles import PatternFill
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
"""
    中国系统 -- 1亿数据分批次导出大量数据到excel,每次一百万
    
    读取两个数据库数据，然后对比， 数据相同的跳过， 数据不同的写入excel
"""

def test_QueryData():  ##查询接口
    wb = openpyxl.Workbook()
    ws = wb.active

    url1 = "http://106.39.185.104:8713/agilorapi/v6/query?db=AGILOR_RTDB"
    url2 = "http://106.39.185.104:8713/agilorapi/v6/query?db=test"

    headers = {
        'Authorization': 'Token XXX',
        'Content-Type': 'application/vnd.agilorql',
        'Accept': 'application/csv'
    }
    limit =10000
    for i in range(0,10000):
        o = limit * i
        data = 'select * from opcua limit ' + str(limit) + ' offset ' + str(o)

        response1 = requests.request("POST",url1, headers=headers, data=data)
        time.sleep(0.5)
        response2 = requests.post(url2,headers=headers,data=data)

        ## 正则处理数据，以\n\t分割数据，返回list
        lista = re.split(r'[\n\t]', response1.text)
        listb = re.split(r'[\n\t]',response2.text)

        if lista == listb:
            print('第%d个10000条数据，备份前与备份后一致' %(i+1))
            print('\n')
        else:
            listrust1 = []
            for k in range(1, len(lista) - 1):
                li1 = []
                li1.append(lista[k])
                listrust1.append(li1)
            print('第%d个lista数据：'%(i+1),lista)

            listrust2 = []
            for h in range(1, len(listb) - 1):
                li2 = []
                li2.append(listb[h])
                listrust2.append(li2)
            print('第%d个listb数据：'%(i+1),listb)

            for u in range(len(listrust1)):
                for g in range(len(listrust1[0])):
                    ws.cell(u + 1, g + 1).value = listrust1[u][g]
            wb.save('E:\sym\全量数据\全量lista0223\第' + str(i+1) + '个' + str(limit) + '条.xlsx')

            for r in range(len(listrust2)):
                for c in range(len(listrust2[0])):
                    ws.cell(r + 1, c + 1).value = listrust2[r][c]
            wb.save('E:\sym\全量数据\全量listb0223\第' + str(i+1) + '个' + str(limit) + '条.xlsx')

            # for r in range(len(listrust1)):
            #     listc = listrust1[r][0].split(',')
            #     for c in range(len(listc)):
            #         ws.cell(r + 1, c + 1).value = listc[c]
            # wb.save('E:\sym\全量数据\全量lista0222\第' + str(i) + '个10000条a.xlsx')
            #
            # for r in range(len(listrust2)):
            #     listc = listrust2[r][0].split(',')
            #     for c in range(len(listc)):
            #         ws.cell(r + 1, c + 1).value = listc[c]
            # wb.save('E:\sym\全量数据\全量listb0222\第' + str(i) + '个10000条b.xlsx')

if __name__ == '__main__':
    total_list = test_QueryData()
















'''
name,       tags,       time,               AGPOINTNAME,        B,          F,                  L,          S
T1,             ,       1642734682916293946,    tag_F_1,         ,          45226.4006899,       ,
T1,             ,       1642734682916293946,    tag_F_10,        ,          79937.8420698,       ,

T1,             ,       1642734683624432481,    tag_L_1,         ,          ,                 44251,
T1,             ,       1642734683624432481,    tag_L_100,       ,          ,                 58713,

T1,             ,       1642734683985583212,    tag_S_1213,       ,         ,                      ,          6e71bbaa-101e-4e11-9d89-5066ccd5ce37
T1,             ,       1642734683985583212,    tag_S_1218,       ,         ,                      ,          3edf901e-2335-46b0-9d1b-e2f5071dd70a

T1,             ,       1642734684305415746,    tag_B_1483,     true,       ,                       ,
T1,             ,       1642734684305415746,    tag_B_1486,     true,       ,                       ,
'''